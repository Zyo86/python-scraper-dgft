import os
import openpyxl
import pandas as pd
from openpyxl import load_workbook, Workbook
import xlrd
from xls2xlsx import XLS2XLSX
import win32com.client as win32

def combine_all(folder_path, type):
    # Set the directory path where the source files are located
    dir_path = f'{folder_path}_xlsx'

    # Create an empty dataframe to hold the collated data
    merged_data = pd.DataFrame()

    # Loop through each file in the directory that has a .xlsx extension
    for filename in os.listdir(dir_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(dir_path, filename)
            
            # Load the Excel file into a pandas dataframe
            data = pd.read_excel(file_path, engine='openpyxl')
            
            # Append the data to the merged_data dataframe
            merged_data = merged_data.append(data, ignore_index=True)

    # Create a new Excel file to hold the merged data
    writer = pd.ExcelWriter(f'collated_data_{type}.xlsx', engine='openpyxl')

    # Write the merged data to the Excel file
    merged_data.to_excel(writer, index=False, engine='openpyxl')

    
    writer.save()


def combine(folder_path, type):
    # create a new workbook to store the collated data
    collated_wb = Workbook()

    # iterate over all the Excel files in the folder
    for file_name in os.listdir(f'{folder_path}_xlsx'):
        # check if the file is an Excel file
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(f'{folder_path}_xlsx', file_name)
            print(f'Processing file: {file_path}')

            # load the Excel file using openpyxl
            wb = load_workbook(filename=file_path)

            # iterate over all the worksheets in the workbook
            for ws in wb.worksheets:
                # iterate over all the rows in the worksheet and append to the collated workbook
                for row in ws.iter_rows(values_only=True):
                    collated_wb.active.append(row)

    # save the collated workbook
    collated_wb.save(os.path.join("", f'collated_{type}.xlsx'))

    print('Collation complete.')

def rename(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith(".xls"):
            newname = filename[:-4] + ".xlsx"
            if not os.path.exists(newname):
                # Load the Excel 97-Excel 2003 workbook using xlrd
                workbook_xls = xlrd.open_workbook(folder_path + '\\' + filename)

                # Create a new Excel workbook using openpyxl
                workbook_xlsx = openpyxl.Workbook()

                # Copy data from the old workbook to the new workbook
                for sheet_name in workbook_xls.sheet_names():
                    worksheet_xls = workbook_xls.sheet_by_name(sheet_name)
                    worksheet_xlsx = workbook_xlsx.create_sheet(title=sheet_name)
                    for row in range(worksheet_xls.nrows):
                        for col in range(worksheet_xls.ncols):
                            cell_value = worksheet_xls.cell_value(row, col)
                            worksheet_xlsx.cell(row=row+1, column=col+1, value=cell_value)

                # Save the new Excel workbook as an .xlsx file
                workbook_xlsx.save(folder_path + '\\' + newname)
                # os.rename(folder_path + '\\' + filename, folder_path + '\\' + newname)
                print(f"Renamed {filename} to {newname}")
            else:
                print(f"Error: {newname} already exists, cannot rename {filename}")


def rename_all(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith(".xls"):
            newname = filename[:-4] + ".xlsx"
            if not os.path.exists(newname):
                x2x = XLS2XLSX(folder_path + '\\' + filename)
                x2x.to_xlsx(folder_path + '\\' + newname)
                print(f"Renamed {filename} to {newname}")
            else:
                print(f"Error: {newname} already exists, cannot rename {filename}")

def saveas_all(folder_path):
    count = 0
    for filename in os.listdir(folder_path):
        print(f"Inside {folder_path}")
        if filename.endswith(".xls"):
            newname = filename[:-4] + ".xlsx"
            if not os.path.exists(newname):
                # Create an instance of the Excel application
                excel = win32.gencache.EnsureDispatch('Excel.Application')

                # Open the .xls file
                workbook = excel.Workbooks.Open(f'{folder_path}\\{filename}')

                # Save the file as .xlsx
                workbook.SaveAs(f'{folder_path}\\{newname}', FileFormat=51)

                # Close the workbook and quit Excel
                workbook.Close()
                excel.Quit()
                print(f"Renamed {filename} to {newname}")
            else:
                print(f"Error: {newname} already exists, cannot rename {filename}")


def formatted_combine(folder_path, type):
    folder_path = f'{folder_path}_xlsx'
    # Open the first file and read the header row
    filename = os.listdir(folder_path)[0]
    workbook = openpyxl.load_workbook(folder_path + '\\' + filename)
    worksheet = workbook.active
    if type == 2:
        header_row = [cell.value for cell in worksheet[9]]
        adjusted = 4
        start_row = 10
    else:
        header_row = [cell.value for cell in worksheet[10]]
        adjusted = 3
        start_row = 11

    # Create a new workbook to store the collated data
    new_workbook = openpyxl.Workbook()
    new_worksheet = new_workbook.active

    # Write the header row to the new worksheet
    for col, header in enumerate(header_row, start=1):
        new_worksheet.cell(row=1, column=col, value=header)

    # Loop through the rest of the files and copy row 10 to the new worksheet
    index = 2
    for filename in os.listdir(folder_path):
        try:
            workbook = openpyxl.load_workbook(folder_path + '\\' + filename)
            worksheet = workbook.active
            max_row = worksheet.max_row
            print("max_row " + str(max_row) + f" - In file {folder_path}\\{filename}")
            adjusted_max_row = max_row - adjusted
            # Loop through each row in the worksheet and copy the data to the new worksheet
            for row in range(start_row, adjusted_max_row + 1):
                row_data = []
                for col in range(1, worksheet.max_column + 1):
                    value = worksheet.cell(row=row, column=col).value
                    row_data.append(value)
                new_worksheet.append(row_data)

            # row_data = [cell.value for cell in worksheet[10]]
            # for col, value in enumerate(row_data, start=1):
            #     new_worksheet.cell(row=index, column=col, value=value)

            index = index + 1
        except FileNotFoundError:
            print(f"{filename} not found")

    # Save the new workbook
    new_workbook.save(f"formatted_{type}.xlsx")
